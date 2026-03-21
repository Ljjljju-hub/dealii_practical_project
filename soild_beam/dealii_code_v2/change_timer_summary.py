import pandas as pd
import re
import os
from datetime import datetime

def parse_performance_file(file_path):
    """
    解析性能分析文件，提取各个section的详细数据
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    
    # 存储解析结果
    section_details = []   # 各section详细数据
    
    current_cycle = None
    current_active_cells = None
    current_dofs = None
    
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        
        # 匹配循环开始
        if "Performance Summary: Cycle" in line:
            # 提取cycle编号
            cycle_match = re.search(r'Cycle (\d+)', line)
            if cycle_match:
                current_cycle = int(cycle_match.group(1))
            
            # 查找Active Cells和DoFs
            j = i + 1
            while j < len(lines) and j < i + 5:
                if "Active Cells:" in lines[j]:
                    active_match = re.search(r'Active Cells:\s+(\d+)', lines[j])
                    if active_match:
                        current_active_cells = int(active_match.group(1))
                
                if "DoFs:" in lines[j]:
                    dofs_match = re.search(r'DoFs:\s+(\d+)', lines[j])
                    if dofs_match:
                        current_dofs = int(dofs_match.group(1))
                j += 1
        
        # 匹配section表格数据
        if line.startswith('|') and 'Section' in line and 'wall time' in line:
            # 跳过表头
            i += 2
            while i < len(lines):
                table_line = lines[i].strip()
                
                # 检查是否到达表格结尾
                if not table_line or table_line.startswith('+---') or 'Total wallclock' in table_line:
                    break
                
                if table_line.startswith('|'):
                    # 解析表格行
                    parts = [p.strip() for p in table_line.split('|')[1:-1]]
                    if len(parts) >= 4:
                        section_name = parts[0].strip()
                        
                        # 跳过空行和表头
                        if section_name and not section_name.startswith('Section'):
                            try:
                                no_calls = int(parts[1]) if parts[1] else 0
                                wall_time = float(parts[2].rstrip('s')) if parts[2] else 0
                                pct_total = float(parts[3].rstrip('%')) if parts[3] else 0
                                
                                section_details.append({
                                    'Cycle': current_cycle,
                                    'Active_Cells': current_active_cells,
                                    'DoFs': current_dofs,
                                    'Section': section_name,
                                    'No_Calls': no_calls,
                                    'Wall_Time_s': wall_time,
                                    'Percent_Total': pct_total
                                })
                            except (ValueError, IndexError):
                                pass
                i += 1
        i += 1
    
    # 转换为DataFrame
    df_details = pd.DataFrame(section_details)
    
    # 排序
    if not df_details.empty:
        df_details = df_details.sort_values(['Cycle', 'Section'])
    
    return df_details

def create_excel_file(df_details, output_file='performance_analysis.xlsx'):
    """
    创建Excel文件，包含Section_Details和Pivot_by_Cycle两个sheet
    """
    if df_details.empty:
        print("警告: 没有数据可保存")
        return
    
    # 创建Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 1. Section_Details sheet - 详细数据表
        df_details.to_excel(writer, sheet_name='Section_Details', index=False)
        format_details_sheet(writer.sheets['Section_Details'], df_details)
        
        # 2. Pivot_by_Cycle sheet - 按循环的透视表（包含Active_Cells和DoFs）
        create_pivot_by_cycle(writer, df_details)
        
        print(f"Excel文件已保存: {output_file}")

def format_details_sheet(worksheet, df):
    """
    格式化详细数据表
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    # 定义样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置列宽
    column_widths = [10, 15, 15, 25, 10, 15, 15]
    headers = ['Cycle', 'Active_Cells', 'DoFs', 'Section', 'No_Calls', 'Wall_Time_s', 'Percent_Total']
    
    for i, (width, header) in enumerate(zip(column_widths, headers), 1):
        worksheet.column_dimensions[get_column_letter(i)].width = width
        # 设置标题
        cell = worksheet.cell(row=1, column=i)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 为不同section设置交替颜色
    colors = ['FFFFFF', 'F2F2F2']
    current_section = None
    color_index = 0
    
    # 格式化数据行
    for row in range(2, len(df) + 2):
        section = worksheet.cell(row=row, column=4).value
        
        # 切换颜色
        if section != current_section:
            current_section = section
            color_index = 1 - color_index
        
        fill_color = colors[color_index]
        
        for col in range(1, 8):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            # 数字格式
            if col == 6:  # Wall_Time_s
                if cell.value and cell.value < 0.001:
                    cell.number_format = '0.000E+00'
                else:
                    cell.number_format = '0.000'
            elif col == 7:  # Percent_Total
                cell.number_format = '0.00'
            elif col in [2, 3]:  # Active_Cells和DoFs
                cell.number_format = '#,##0'
    
    # 添加筛选
    worksheet.auto_filter.ref = f'A1:G{len(df)+1}'
    
    # 冻结首行
    worksheet.freeze_panes = 'A2'

def create_pivot_by_cycle(writer, df_details):
    """
    创建按循环的透视表，包含Active_Cells和DoFs信息
    """
    workbook = writer.book
    worksheet = workbook.create_sheet('Pivot_by_Cycle')
    
    if df_details.empty:
        worksheet['A1'] = '无数据可分析'
        return
    
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # 获取所有唯一的Cycle和Section
    cycles = sorted(df_details['Cycle'].unique())
    sections = sorted(df_details['Section'].unique())
    
    # 为每个Cycle获取Active_Cells和DoFs
    cycle_info = {}
    for cycle in cycles:
        cycle_data = df_details[df_details['Cycle'] == cycle]
        if not cycle_data.empty:
            cycle_info[cycle] = {
                'Active_Cells': cycle_data['Active_Cells'].iloc[0],
                'DoFs': cycle_data['DoFs'].iloc[0]
            }
    
    # 创建透视表数据
    pivot_data = pd.pivot_table(
        df_details,
        values='Wall_Time_s',
        index='Section',
        columns='Cycle',
        aggfunc='sum',
        fill_value=0
    )
    
    # 添加总计行
    pivot_data.loc['总计'] = pivot_data.sum()
    
    # 写入数据
    current_row = 1
    
    # 写入标题行（Cycle信息）
    worksheet.cell(row=current_row, column=1, value='Section / Cycle')
    for c_idx, cycle in enumerate(cycles, start=2):
        worksheet.cell(row=current_row, column=c_idx, value=f'Cycle {cycle}')
    
    # 设置标题行样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, len(cycles) + 2):
        cell = worksheet.cell(row=current_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    current_row += 1
    
    # 写入Active_Cells行
    worksheet.cell(row=current_row, column=1, value='Active_Cells')
    for c_idx, cycle in enumerate(cycles, start=2):
        worksheet.cell(row=current_row, column=c_idx, value=cycle_info[cycle]['Active_Cells'])
    
    # 设置Active_Cells行样式
    info_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col in range(1, len(cycles) + 2):
        cell = worksheet.cell(row=current_row, column=col)
        cell.fill = info_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        if col > 1:
            cell.number_format = '#,##0'
    
    current_row += 1
    
    # 写入DoFs行
    worksheet.cell(row=current_row, column=1, value='DoFs')
    for c_idx, cycle in enumerate(cycles, start=2):
        worksheet.cell(row=current_row, column=c_idx, value=cycle_info[cycle]['DoFs'])
    
    # 设置DoFs行样式
    for col in range(1, len(cycles) + 2):
        cell = worksheet.cell(row=current_row, column=col)
        cell.fill = info_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        if col > 1:
            cell.number_format = '#,##0'
    
    current_row += 1
    
    # 添加空行
    current_row += 1
    
    # 写入Section数据
    for r_idx, (section, row_data) in enumerate(pivot_data.iterrows(), start=current_row):
        worksheet.cell(row=r_idx, column=1, value=section)
        for c_idx, value in enumerate(row_data, start=2):
            worksheet.cell(row=r_idx, column=c_idx, value=value)
    
    # 设置Section数据区域样式
    for row in range(current_row, current_row + len(pivot_data)):
        for col in range(1, len(cycles) + 2):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if col == 1:  # Section名称列
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            else:  # 数值列
                if cell.value:
                    if cell.value < 0.001:
                        cell.number_format = '0.000E+00'
                    else:
                        cell.number_format = '0.000'
    
    # 高亮总计行
    total_row = current_row + len(pivot_data) - 1
    for col in range(1, len(cycles) + 2):
        cell = worksheet.cell(row=total_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    
    # 调整列宽
    worksheet.column_dimensions['A'].width = 25
    for col in range(2, len(cycles) + 2):
        worksheet.column_dimensions[chr(64 + col)].width = 15

def main():
    # 文件路径
    input_file = './build/timer_summary.txt'
    output_file = 'performance_analysis.xlsx'
    
    try:
        # 解析文件
        print("正在解析文件...")
        df_details = parse_performance_file(input_file)
        
        # 打印统计信息
        print(f"\n解析完成！")
        print(f"详细数据: {len(df_details)} 条记录")
        
        if not df_details.empty:
            cycles = sorted(df_details['Cycle'].unique())
            sections = sorted(df_details['Section'].unique())
            print(f"包含的Cycle: {cycles}")
            print(f"包含的Section: {sections}")
            
            # 显示数据预览
            print("\n数据预览（前5条）:")
            print("="*80)
            preview_cols = ['Cycle', 'Active_Cells', 'DoFs', 'Section', 'Wall_Time_s', 'Percent_Total']
            print(df_details[preview_cols].head(10).to_string(index=False))
            print("="*80)
        else:
            print("警告: 没有找到任何数据")
        
        # 创建Excel文件
        print(f"\n正在生成Excel文件...")
        create_excel_file(df_details, output_file)
        
        print(f"\n处理完成！")
        print(f"生成的文件: {output_file}")
        print(f"包含的Sheet:")
        print(f"  - Section_Details: 详细数据表")
        print(f"  - Pivot_by_Cycle: 按循环的透视表（包含网格数和自由度信息）")
        
    except FileNotFoundError:
        print(f"错误: 找不到文件 {input_file}")
        print(f"当前目录: {os.getcwd()}")
    except Exception as e:
        print(f"处理过程中出现错误: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()